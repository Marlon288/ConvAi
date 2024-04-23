import json
import openpyxl
import os


json_file_path = "api/data/prompts.json"

# Get the absolute path of the Python script
script_directory = os.path.dirname(os.path.abspath(__file__))

# Combine the script directory path with the relative JSON file path
json_file_path = os.path.join(script_directory, json_file_path)

# Load the JSON data from the file
with open(json_file_path) as file:
    data = json.load(file)


# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet_gpt4 = workbook.active
sheet_gpt4.title = "GPT-4"

sheet_gpt4.append(["LLM", "Performance"])


sheet_gpt35 = workbook.create_sheet(title="GPT-3.5")


sheet_gpt35.append(["LLM", "Performance"])


gpt4_sum = 0
gpt4_count = 0
gpt35_sum = 0
gpt35_count = 0


for item in data:
    llm = item["LLM"]
    performance = item["Performance"]

    if llm == "gpt-4-1106-preview":
        sheet_gpt4.append([llm, performance])
        gpt4_sum += performance
        gpt4_count += 1
    elif llm == "gpt-3.5-turbo":
        sheet_gpt35.append([llm, performance])
        gpt35_sum += performance
        gpt35_count += 1


if gpt4_count > 0:
    gpt4_average = gpt4_sum / gpt4_count
    sheet_gpt4.insert_rows(1)
    sheet_gpt4.cell(row=1, column=1, value="Average Performance")
    sheet_gpt4.cell(row=1, column=2, value=gpt4_average)


if gpt35_count > 0:
    gpt35_average = gpt35_sum / gpt35_count
    sheet_gpt35.insert_rows(1)
    sheet_gpt35.cell(row=1, column=1, value="Average Performance")
    sheet_gpt35.cell(row=1, column=2, value=gpt35_average)


workbook.save("performance_data.xlsx")